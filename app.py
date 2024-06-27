import os
import json
import secrets
from io import BytesIO

from bson import ObjectId
from flask import Flask, render_template, request, redirect, url_for, session, flash, make_response
from flask_pymongo import PyMongo
from twilio.rest import Client
import re

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd
from werkzeug.utils import secure_filename
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import datetime

app = Flask(__name__)

# Генерация случайного ключа
app.secret_key = os.urandom(24)

# MongoDB подключение
app.config['MONGO_URI'] = 'mongodb://localhost:27017/RequestsProj'

# MongoDB
mongo = PyMongo(app)

# группа пользователей
ADMINS = "ADMINS"
REQUESTS = "REQUESTS"
adminsTypes = ["Главный", "Не главный"]

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


sender_email = "email"
sender_password = "password"
smtp_server = "smtp.timeweb.ru"
smtp_port = 465
# Настройка SMTP сессии
server = smtplib.SMTP_SSL(smtp_server, smtp_port)
server.login(sender_email, sender_password)  # Вход в аккаунт
def send_email( recipient, body):
    # Создание MIME сообщения
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient
    msg['Subject'] = "Wifi сеть"

    # Добавление тела письма
    msg.attach(MIMEText(body, 'plain'))

    try:


        # Отправка письма
        text = msg.as_string()
        server.sendmail(sender_email, recipient, text)

        # Закрытие соединения
        #server.quit()

        print("Email sent successfully!")
    except Exception as e:
        print(f"Failed to send email: {e}")

def log(message):
    with open('logs/logs.txt', 'a') as file:
        user = mongo.db[ADMINS].find_one({'admin_login': session.get('username')})
        current_datetime = datetime.datetime.now()
        formatted_datetime = current_datetime.strftime('%Y-%m-%d %H:%M:%S')
        file.write(f'{formatted_datetime} - {message}  админ: {user.get('admin_last_name')} {user.get('admin_first_name')} {user.get('admin_middle_name')}\n')


def check_duplicate_record(record):
    # Проверка наличия дубликата записи в базе данных
    query = {
        'phone': record['phone'],
        'request_number': record['request_number'],
        'login': record['login'],
        'password': record['password']
    }
    existing_record = mongo.db[REQUESTS].find_one(query)
    return existing_record is not None

def check_duplicate_admin(admin):
    # Проверка наличия дубликата записи в базе данных
    query = {
        'admin_login': admin['admin_login'],
        'admin_password': admin['admin_password']
    }
    existing_record = mongo.db[ADMINS].find_one(query)
    return existing_record is not None



@app.route('/')
def login():
    return render_template('index.html')

@app.route('/check_login', methods=['POST'])
def check_login():
    admin_login = request.form['username']
    password = request.form['password']

    user = mongo.db[ADMINS].find_one({'admin_login': admin_login})

    if user and user['admin_password'] == password:
        session['username'] = user['admin_login']
        if user.get('admin_type') == adminsTypes[1]:
            return redirect(url_for('dashboard'))
        elif user.get('admin_type') == adminsTypes[0]:
            return redirect(url_for('dashboardAdmin'))
    else:
        return render_template('index.html', error_message="Вход не удался! Пожалуйста, проверьте свое имя пользователя и пароль.")



@app.route('/dashboardAdmin', methods=['GET'])
def dashboardAdmin():

    username = session.get('username')
    if not username:
        return redirect(url_for('login'))

    generated_password = secrets.token_urlsafe(12)
    admin_generated_password = secrets.token_urlsafe(12)

    # Получаем результаты поиска из сессии
    searched_requests = session.get('searched_requests')
    searched_admins = session.get('searched_admins')

    if searched_requests is not None:
        user_requests = searched_requests
    else:
        # Иначе загружаем все заявки из базы данных
        user_requests = list(mongo.db[REQUESTS].find({}))

    if searched_admins is not None:
        admins = searched_admins
    else:
        # Иначе загружаем всех администраторов из базы данных
        admins = list(mongo.db[ADMINS].find({}))
    search_query = ""

    if session.get('searched_search_query') is not None:
        search_query = session.get('searched_search_query')
    return render_template('dashboardAdmin.html',
                           username=username,
                           user_requests=user_requests,
                           admins=admins,
                           generated_password=generated_password,
                           admin_generated_password=admin_generated_password,
                           search_query=search_query)
@app.route('/dashboard', methods=['GET'])
def dashboard():
    username = session.get('username')
    if not username:
        return redirect(url_for('login'))

    generated_password = secrets.token_urlsafe(12)
    search_query=""
    searched_requests = session.get('searched_requests')
    if session.get('searched_search_query') is not None:
        search_query = session.get('searched_search_query')
    if searched_requests is not None:
        user_requests = searched_requests
    else:
        user_requests = list(mongo.db[REQUESTS].find({}))

    return render_template('dashboard.html',
                           username=username,
                           user_requests=user_requests,
                           generated_password=generated_password,
                           search_query=search_query)


@app.route('/search_requests/<arg1>', methods=['POST'])
def search_requests(arg1):
    search_query_string = request.form.get('search_query', '')
    regex = re.compile(search_query_string, re.IGNORECASE)
    results = list(mongo.db[REQUESTS].find({'last_name': {'$regex': regex}}))
    print(results)
    # Преобразуем ObjectId в строки для сохранения в сессию
    for result in results:
        result['_id'] = str(result['_id'])

    # Сохраняем результаты поиска в сессию
    session['searched_requests'] = results
    session['searched_search_query'] =request.form.get('search_query', '')
    return redirect(url_for(arg1))


@app.route('/add_request/<arg1>/<arg2>', methods=['POST'])
def add_request(arg1,arg2):
    if request.method == 'POST':


        user = mongo.db[ADMINS].find_one({'admin_login': session.get('username')})

        # Формирование данных для добавления
        record = {
            'request_number': request.form['request_number'],
            'last_name': request.form['last_name'],
            'first_name': request.form['first_name'],
            'middle_name': request.form['middle_name'],
            'phone': request.form['phone'],
            'email': request.form['email'],
            'ssid': request.form['ssid'],
            'login': request.form['login'],
            'password': request.form['password'],
            'message': f"Ваш персональный пароль для подключения к сети {request.form['ssid']} : {arg1}",
            'adminInitials': f"{user.get('admin_last_name')} {user.get('admin_first_name')} {user.get('admin_middle_name')}"

        }

            # Добавление заявки в базу данных
        if check_duplicate_record(record):
            return "Такая запись уже существует в базе данных."

        try:
            # Добавление заявки в базу данных
            mongo.db[REQUESTS].insert_one(record)
            log(f'Была добавлена заявка - {record.get("request_number")}')
            return redirect(url_for(arg2))
        except Exception as e:
            return ({'status': 'error', 'message': f'Ошибка при добавлении записи в базу данных: {str(e)}'})


@app.route('/add_admin/', methods=['POST'])
def add_admin():
    if request.method == 'POST':
        if 'admin_last_name' in request.form:
            # Проверка на пустоту полей
            required_fields = ['admin_last_name', 'admin_first_name', 'admin_middle_name', 'admin_type', 'admin_login', 'admin_password']
            for field in required_fields:
                if not request.form.get(field):
                    flash(f'Поле {field} не должно быть пустым', 'danger')
                    return redirect(url_for('dashboardAdmin'))


            admin = {
                'admin_last_name': request.form['admin_last_name'],
                'admin_first_name': request.form['admin_first_name'],
                'admin_middle_name': request.form['admin_middle_name'],
                'admin_type': request.form['admin_type'],
                'admin_login': request.form['admin_login'],
                'admin_password': request.form['admin_password']
            }
            # Добавление администратора в базу данных
            if check_duplicate_admin(admin):
                return "Такой админ уже существует в базе данных."

            try:
                # Добавление заявки в базу данных
                mongo.db[ADMINS].insert_one(admin)
                log(f'Была добавлен админ - {admin.get('admin_last_name')} {admin.get('admin_first_name')} {admin.get('admin_middle_name')}')

            except Exception as e:
                return ({'status': 'error', 'message': f'Ошибка при добавлении админа в базу данных: {str(e)}'})




        return redirect(url_for('dashboardAdmin'))


@app.route('/delete_admins', methods=['POST'])
def delete_admins():
    try:
        selected_admins = request.form.getlist('selected_admins')
        print("Selected Admins:", selected_admins)

        # Найти текущего администратора
        current_admin = mongo.db[ADMINS].find_one({'admin_login': session.get('username')})
        current_admin_id = str(current_admin['_id'])
        print("Current Admin ID:", current_admin_id)

        # Удалить текущего администратора из списка для удаления, если он есть
        if current_admin_id in selected_admins:
            selected_admins.remove(current_admin_id)
            print("Selected Admins after removing current admin:", selected_admins)

        # Проверка и преобразование идентификаторов в ObjectId
        valid_object_ids = [ObjectId(admin_id) for admin_id in selected_admins if ObjectId.is_valid(admin_id)]
        print("Valid ObjectIds:", valid_object_ids)

        if valid_object_ids:
            for obj_id in valid_object_ids:
                temp = mongo.db[ADMINS].find_one({'_id': obj_id})
                print(temp)
                if temp:
                    log(f'Был удален администратор - {temp.get('admin_last_name')} {temp.get('admin_first_name')} {temp.get('admin_middle_name')}')
                    mongo.db[ADMINS].delete_one(temp)
                else:
                    print(f"Document not found for id: {obj_id}")

        return redirect(url_for('dashboardAdmin'))

    except Exception as e:
        print(f"Error while deleting admins: {e}")
        return redirect(url_for('dashboardAdmin'))

@app.route('/delete_requests/<arg1>', methods=['POST'])
def delete_requests(arg1):
    try:
        session.pop('searched_requests', None)
        session.pop('searched_search_query', None)
        selected_requests = request.form.getlist('selected_requests')
        print("Selected Requests:", selected_requests)

        # Проверка и преобразование идентификаторов в ObjectId
        valid_object_ids = [ObjectId(req_id) for req_id in selected_requests if ObjectId.is_valid(req_id)]
        print("Valid ObjectIds:", valid_object_ids)

        if valid_object_ids:
            for obj_id in valid_object_ids:
                temp = mongo.db[REQUESTS].find_one({'_id': obj_id})
                if temp:
                    log(f'Была удалена заявка - {temp.get("request_number")}')
                    mongo.db[REQUESTS].delete_one(temp)
                else:
                    print(f"Document not found for id: {obj_id}")

        return redirect(url_for(arg1))

    except Exception as e:
        print(f"Error while deleting requests: {e}")
        return redirect(url_for(arg1))


@app.route('/import_excel/<arg1>', methods=['POST'])
def import_excel(arg1):
    try:
        if 'file' not in request.files:
            return "Файл не найден."

        file = request.files['file']
        if file.filename == '':
            return "Файл не выбран."

        if file and file.filename.endswith('.xlsx'):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)

            # Чтение Excel файла
            wb = load_workbook(file_path)
            sheet = wb.active
            filled_rows = sum(1 for row in sheet.iter_rows(min_row=2) if any(cell.value for cell in row))

            for i in range(2, filled_rows + 2):
                try:
                    # Получаем данные из Excel строки
                    last_name, first_name, middle_name = str(sheet[f'A{i}'].value).split(' ', 2)

                    user = mongo.db[ADMINS].find_one({'admin_login': session.get('username')})

                    # Формируем запись для MongoDB
                    record = {
                        'request_number': str(sheet[f'B{i}'].value),
                        'last_name': last_name,
                        'first_name': first_name,
                        'middle_name': middle_name,
                        'phone': str(sheet[f'C{i}'].value),
                        'email': str(sheet[f'D{i}'].value),
                        'ssid': str(sheet[f'E{i}'].value),
                        'login': str(sheet[f'F{i}'].value),
                        'password': str(sheet[f'G{i}'].value),
                        'message': f"Ваш персональный пароль для подключения к сети {sheet[f'D{i}'].value} : {str(sheet[f'G{i}'].value)}",
                        'adminInitials': f"{user.get('admin_last_name')} {user.get('admin_first_name')} {user.get('admin_middle_name')}"
                    }

                    # Вставляем данные в MongoDB
                    if not check_duplicate_record(record):
                        mongo.db[REQUESTS].insert_one(record)
                        log(f'Была импортирована заявка - {record.get("request_number")}')

                    else:
                        continue


                except Exception as e:
                    print(f"Ошибка при формировании записи для строки {i}: {e}")
                    continue

            wb.close()
            return redirect(url_for(arg1))

        else:
            return "Неверный формат файла. Пожалуйста, загрузите .xlsx файл."

    except Exception as e:
        print(f"Ошибка при импорте Excel файла: {e}")
        return "Произошла ошибка при импорте Excel файла."


@app.route('/export_excel/<arg1>', methods=['GET'])
def export_excel(arg1):
    try:
        # Создаем новый Excel файл
        wb = Workbook()
        sheet = wb.active

        # Задаем заголовки столбцов
        sheet[f'A{1}'].value="ФИО"
        sheet[f'B{1}'].value="Заявка в SD"
        sheet[f'C{1}'].value="Телефон"
        sheet[f'D{1}'].value="Почта"
        sheet[f'E{1}'].value="SSID"
        sheet[f'F{1}'].value="Логин"
        sheet[f'G{1}'].value="Пароль"
        sheet[f'H{1}'].value="Текст СМС"

        # Получаем данные из MongoDB
        records = mongo.db[REQUESTS].find()

        for record in records:
            # Формируем строку данных
            row = [
                f"{record.get('last_name')} {record.get('first_name')} {record.get('middle_name')}",
                record.get('request_number'),
                record.get('phone'),
                record.get('email'),
                record.get('ssid'),
                record.get('login'),
                record.get('password'),
                record.get('message')
            ]
            sheet.append(row)

        # Сохраняем файл в объект BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename="Zayavki.xlsx"'
        return response

    except Exception as e:
        print(f"Ошибка при экспорте данных в Excel файл: {e}")
        return "Произошла ошибка при экспорте данных в Excel файл."


@app.route('/send_messages/<arg1>', methods=['POST'])
def send_messages(arg1):
    try:
        selected_requests_json = request.form.get('selected_requests')

        if selected_requests_json:
            selected_requests = json.loads(selected_requests_json)
            print("Selected Requests:", selected_requests)

            valid_object_ids = [ObjectId(req_id) for req_id in selected_requests if ObjectId.is_valid(req_id)]
            print("Valid ObjectIds:", valid_object_ids)

            method = request.form.get("send_method")
            if valid_object_ids:
                for obj_id in valid_object_ids:
                    print(obj_id)
                    temp = mongo.db[REQUESTS].find_one({'_id': obj_id})
                    print(temp)
                    if temp:
                        print(method)
                        if method == "email":
                            to_email = temp.get('email')
                            if to_email !="None":
                                email_body = temp.get('message')
                                send_email(to_email,  email_body)
                                print("Sending email")
                                log(f'Было отправдено сообщение на почту по заявке - {temp.get("request_number")}')
                        elif method == "phone":
                            print("Sending SMS")
                            log(f'Было отправдено сообщение на мобильный телефон по заявке - {temp.get("request_number")}')
                    else:
                        print(f"Document not found for id: {obj_id}")

            return redirect(url_for(arg1))
        else:
            raise ValueError("No selected_requests found in form data")

    except Exception as e:
        print(f"Error while sending messages: {e}")
        return redirect(url_for(arg1))

if __name__ == '__main__':
    app.run()
